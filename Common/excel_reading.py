#!/usr/bin/python3
# -*- coding:utf-8 -*-
# @Author:tan
# @Time:2019/10/11 10:22
# @Email:1164355091@qq.com
import openpyxl
#封装一个读写用例的Excel类
#实现功能
#1.读取用例数据
#2.写入数据

class Excel():
    #实例对象生成时执行  1.打开文件，生成工作簿  2.获取操作的表单
    def __init__(self,wb,sheetname):
        '''
        :param wb: 文件名
        :param sheetname:表单名
        '''
        self.file=wb
        self.wb=openpyxl.load_workbook(wb)
        self.sheetname=self.wb[sheetname]
    #实例对象消亡后执行  关闭工作簿
    def __del__(self):
        self.wb.close()
    #读取单元格中的内容
    def cell_read(self,ro,col):
        '''
        :param ro:指定行
        :param col: 指定列
        :return: cell_read():读取出指定单元格中的内容
        '''
        cell_value=self.sheetname.cell(row=ro,column=col).value  #读取出所指定的单元格中的内容
        return cell_value
    #读取表中所有内容
    def sheet_read(self):
        '''
        :param max_row是所选表中的所有行数
        :param max_column是所选表中的所有列数
        :return:sheet_read():读取所选表中的所有内容按照列表类型表现出来，其中元素以元祖类型表现
        '''
        max_row=self.sheetname.max_row
        max_column=self.sheetname.max_column
        hang_list=[]       #创建一个空列表来储存单行的内容
        zonghang_list=[]   #创建一个空列表用来储存所有行的内容
        for i in range(1,max_row+1):   #遍历出所有行
            for j in range(1,max_column+1):   #遍历出所有列
                values=self.sheetname.cell(row=i,column=j).value
                hang_list.append(str(values))      #将单行的内容以字符串类型添加到单行列表中
                for x in range(0,len(hang_list)):  #将单行列表遍历出来
                    try:            #eval()中是中文汉字的情况下会出现报错，使用异常处理过渡报错
                        if type(eval(hang_list[x]))==dict or type(eval(hang_list[x]))==tuple or type(eval(hang_list[x]))==list:  #当列表中元素为字典、元祖和列表类型时，使用eval()返回元素原先的序列类型
                            hang_list[x]=eval(hang_list[x])
                    except Exception:
                        pass
            zonghang_list.append(tuple(hang_list)) #将单行列表以元祖类型添加到总行列表中
            hang_list.clear()
        return zonghang_list
    #读取指定行
    def row_appointed(self,row_number):
        '''
        :param row_number: 指定行
        :param max_column: 所有列数
        :return: 读取出的指定行所有列的内容以列表类型表现出来
        '''
        self.row_number=row_number
        max_column=self.sheetname.max_column  #获取所有列数
        row_appointed_values=[]  #创建一个空列表用来存储指定行的内容
        for i in range(1,max_column+1):
            row_appointed_values.append(str(self.sheetname.cell(row=row_number,column=i).value))   #将指定行所有列的内容以字符串类型添加到指定行列表中
        return row_appointed_values
    #读取指定列的内容
    def col_appointed(self,col_number):
        '''
        :param col_number: 指定列
        :param max_row: 所有行数
        :return: 读取出的指定列的内容以列表类型表现出来
        '''
        self.col_number=col_number
        max_row=self.sheetname.max_row  #获取所有行数
        col_appointed_values=[]    #创建一个空列表用来存储指定列的内容
        for i in range(1,max_row+1):   #遍历出所有行
            col_appointed_values.append(str(self.sheetname.cell(row=i,column=col_number).value))   #将指定列所有行的内容以字符串类型添加到指定列列表中
        return col_appointed_values
    #读取所需行所需列的内容以元素为元祖类型表现出来
    def row_col_appointed_tuple(self,row_number,col_number):      # row_number,col_number为多行多列时，需要写成列表类型
        self.row_number=list(row_number)  #所需行，需要写成列表类型
        self.col_number=list(col_number)  #所需列，需要写成列表类型
        row_col_appointed_values=[]       #创建一个空列表用来存储所需行所需列的内容
        for i in self.row_number:         #遍历出所需行
            hang_list=[]                  #创建一个空列表用来临时存储单行的内容，每次循环会清空
            for j in self.col_number:     #遍历出所需列
                hang_list.append(str(self.sheetname.cell(row=i,column=j).value))     #将单行的内容以字符串类型添加到单行列表中
                for x in range(0,len(hang_list)):  #将单行列表遍历出来
                    try:            #eval()中是中文汉字的情况下会出现报错，使用异常处理过渡报错
                        if type(eval(hang_list[x]))==dict or type(eval(hang_list[x]))==tuple or type(eval(hang_list[x]))==list:  #当列表中元素为字典、元祖和列表类型时，使用eval()返回元素原先的序列类型
                            hang_list[x]=eval(hang_list[x])
                    except Exception:
                        pass
            row_col_appointed_values.append(tuple(hang_list))         #将单行列表以元祖类型添加到总行列表中
        return row_col_appointed_values
    #写入数据
    def write_data(self,row_number,column_number,msg):
        '''

        :param row_number: 指定行
        :param column_number: 指定列
        :param msg: 所填入的内容
        :return:在指定位置写入内容
        '''
        self.sheetname.cell(row=row_number,column=column_number,value=msg)
        self.wb.save(self.file)  #保存

    def max_column(self):
        return self.sheetname.max_column
    def write_data_lineskip(self,row_number,col_number,msg):
        max_row=self.sheetname.max_row
        for i in range(row_number,max_row+1):
            self.sheetname.cell(row=i, column=col_number, value=msg)
            self.wb.save(self.file)  # 保存

        # 读取所需行所需列的内容
    def row_col_appointed(self, row_number, col_number):  # row_number为多行时，需要写成列表类型
        self.row_number = list(row_number)  # 所需行，需要写成列表类型
        self.col_number =col_number  # 所需列，需要写成整数类型
        hang_list = [] # 创建一个空列表用来存储所需的内容
        for i in self.row_number:  # 遍历出所需行
            hang_list.append(str(self.sheetname.cell(row=i, column=self.col_number).value))  # 将内容以字符串类型添加到所需列表中
            for x in range(0, len(hang_list)):  # 将所需列表遍历出来
                try:  # eval()中是中文汉字的情况下会出现报错，使用异常处理过渡报错
                    if type(eval(hang_list[x])) == dict or type(eval(hang_list[x])) == tuple or type(
                            eval(hang_list[x])) == list:  # 当列表中元素为字典、元祖和列表类型时，使用eval()返回元素原先的序列类型
                        hang_list[x] = eval(hang_list[x])
                except Exception:
                    pass
        return hang_list



# if __name__=="__main__":
#     a=Excel(r"E:\python代码\Unittest_demo\Data\test.xlsx","Sheet")
#     print(a.cell_read(2,3))
#     print(a.sheet_read())
#     print(a.row_appointed(2))
#     print(a.col_appointed(3))
#     print(a.row_col_appointed([3,4],[2,3,4]))
#     print(a.write_data(2,3,"over"))
#     print(a.write_data_lineskip(1,7,"645654"))

